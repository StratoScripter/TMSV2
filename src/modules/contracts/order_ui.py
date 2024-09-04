from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, 
                               QTableWidgetItem, QHeaderView, QLineEdit, QMessageBox, QDialog,
                               QFormLayout, QComboBox, QDoubleSpinBox, QDateTimeEdit, QTextEdit,
                               QAbstractItemView, QFileDialog)
from PySide6.QtCore import Qt, Slot
from PySide6.QtGui import QIcon, QTextDocument
from PySide6.QtPrintSupport import QPrinter, QPrintDialog
from .order import order_module
from modules.product.server import server_module
from ui.ui_styles import apply_styles, Styles, TMSWidget
import logging
import openpyxl
from typing import Optional, Dict, Any

logger = logging.getLogger(__name__)

class OrderWidget(TMSWidget):
    def __init__(self):
        super().__init__()
        self.is_connected = False
        apply_styles(self)
        self.init_ui()
        order_module.data_updated.connect(self.refresh_table)
        order_module.error_occurred.connect(self.show_error_message)
        server_module.connection_status_changed.connect(self.set_connection_status)

    def show_error_message(self, message: str):
        logger.error(message)
        QMessageBox.critical(self, "Error", message)
    
    def show_info_message(self, message: str):
        logger.info(message)
        QMessageBox.information(self, "Information", message)

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # Search bar
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search Orders...")
        
        self.search_button = QPushButton("Search")
        
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_button)
        layout.addLayout(search_layout)

        # Buttons
        button_layout = QHBoxLayout()
        self.add_button = QPushButton("Add Order")
        self.edit_button = QPushButton("Edit Order")
        self.delete_button = QPushButton("Delete Order")
        self.refresh_button = QPushButton("Refresh")
        self.print_button = QPushButton("Print")
        self.export_button = QPushButton("Export to Excel")
        button_layout.addWidget(self.print_button)
        button_layout.addWidget(self.export_button)

        for button in [self.add_button, self.edit_button, self.delete_button, self.refresh_button]:
            
            button_layout.addWidget(button)

        layout.addLayout(button_layout)

        # Table
        self.order_table = QTableWidget()
        
        self.order_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.order_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.order_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        layout.addWidget(self.order_table)

        # Connect signals
        self.search_button.clicked.connect(self.search_orders)
        self.add_button.clicked.connect(self.add_order)
        self.edit_button.clicked.connect(self.edit_order)
        self.delete_button.clicked.connect(self.delete_order)
        self.refresh_button.clicked.connect(self.refresh_table)
        self.print_button.clicked.connect(self.print_order)
        self.export_button.clicked.connect(self.export_to_excel)

        self.update_ui_elements()

    @Slot(bool, str)
    def set_connection_status(self, is_connected: bool, message: str):
        self.is_connected = is_connected
        self.update_ui_elements()
        if is_connected:
            self.refresh_table()
        else:
            self.order_table.setRowCount(0)

    def update_ui_elements(self):
        enabled = self.is_connected
        for widget in [self.add_button, self.edit_button, self.delete_button,
                       self.refresh_button, self.search_button, self.search_input]:
            widget.setEnabled(enabled)

    def refresh_table(self):
        if not self.is_connected:
            return

        self.order_table.clear()
        if isinstance(order_module.orders, list):
            self.order_table.setRowCount(len(order_module.orders))
            self.order_table.setColumnCount(19)
            self.order_table.setHorizontalHeaderLabels([
                "Order ID", "Seller", "Buyer", "Product", "Driver", "Vehicle",
                "Loading Arm", "Storage Tank", "Weighbridge", "Planned Quantity",
                "Actual Quantity", "Unit", "Initial Weight", "Final Weight",
                "Checkout Weight", "Order Date", "Checkout Time", "Status", "Remarks"
            ])

            for row, order in enumerate(order_module.orders):
                self.order_table.setItem(row, 0, QTableWidgetItem(str(order.get('OrderId', ''))))
                self.order_table.setItem(row, 1, QTableWidgetItem(order.get('SellerName', '')))
                self.order_table.setItem(row, 2, QTableWidgetItem(order.get('BuyerName', '')))
                self.order_table.setItem(row, 3, QTableWidgetItem(order.get('ProductName', '')))
                self.order_table.setItem(row, 4, QTableWidgetItem(order.get('DriverName', '')))
                self.order_table.setItem(row, 5, QTableWidgetItem(order.get('LicensePlateNumber', '')))
                self.order_table.setItem(row, 6, QTableWidgetItem(order.get('LoadingArmName', '')))
                self.order_table.setItem(row, 7, QTableWidgetItem(order.get('StorageTankName', '')))
                self.order_table.setItem(row, 8, QTableWidgetItem(order.get('WeighbridgeName', '')))
                self.order_table.setItem(row, 9, QTableWidgetItem(str(order.get('PlannedQuantity', ''))))
                self.order_table.setItem(row, 10, QTableWidgetItem(str(order.get('ActualQuantity', ''))))
                self.order_table.setItem(row, 11, QTableWidgetItem(order.get('UnitName', '')))
                self.order_table.setItem(row, 12, QTableWidgetItem(str(order.get('InitialWeight', ''))))
                self.order_table.setItem(row, 13, QTableWidgetItem(str(order.get('FinalWeight', ''))))
                self.order_table.setItem(row, 14, QTableWidgetItem(str(order.get('CheckoutWeight', ''))))
                self.order_table.setItem(row, 15, QTableWidgetItem(str(order.get('LoadingDate', ''))))
                self.order_table.setItem(row, 16, QTableWidgetItem(str(order.get('LoadingTime', ''))))
                self.order_table.setItem(row, 17, QTableWidgetItem(order.get('Status', '')))
                self.order_table.setItem(row, 18, QTableWidgetItem(order.get('Remarks', '')))

            self.order_table.resizeColumnsToContents()
            self.order_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        else:
            logger.error(f"Unexpected type for order_module.orders: {type(order_module.orders)}")

    def search_orders(self):
        if not self.is_connected:
            return
        search_term = self.search_input.text().strip()
        order_module.search_orders(search_term)

    def add_order(self):
        if not self.is_connected:
            return
        dialog = OrderDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            success, order_id, message = order_module.add_order(dialog.get_data())
            if success:
                logger.info(f"Order added successfully. ID: {order_id}")
                self.refresh_table()
            else:
                self.show_error_message(f"Failed to add order: {message}")

    def edit_order(self):
        if not self.is_connected:
            return
        selected_row = self.order_table.currentRow()
        if 0 <= selected_row < len(order_module.orders):
            order_id = int(self.order_table.item(selected_row, 0).text())
            dialog = OrderDialog(self, order_id=order_id)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                success, message = order_module.update_order(dialog.get_data())
                if success:
                    logger.info(f"Order {order_id} updated successfully.")
                    self.refresh_table()
                else:
                    self.show_error_message(f"Failed to update order: {message}")
        else:
            self.show_error_message("Please select an order to edit.")

    def delete_order(self):
        if not self.is_connected:
            return
        selected_row = self.order_table.currentRow()
        if 0 <= selected_row < len(order_module.orders):
            order_id = int(self.order_table.item(selected_row, 0).text())
            confirm = QMessageBox.question(self, "Confirm Deletion", "Are you sure you want to delete this order?",
                                           QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if confirm == QMessageBox.StandardButton.Yes:
                if order_module.delete_order(order_id):
                    logger.info(f"Order {order_id} deleted successfully.")
                    self.refresh_table()
                else:
                    self.show_error_message("Failed to delete the order.")
        else:
            self.show_error_message("Please select an order to delete.")

    def print_order(self):
        if not self.is_connected:
            return
        selected_row = self.order_table.currentRow()
        if 0 <= selected_row < len(order_module.orders):
            order_id = int(self.order_table.item(selected_row, 0).text())
            order_data = order_module.get_order_details(order_id)
            if order_data:
                printer = QPrinter(QPrinter.PrinterMode.HighResolution)
                dialog = QPrintDialog(printer, self)
                if dialog.exec() == QDialog.DialogCode.Accepted:
                    self.print_document(printer, order_data)
        else:
            self.show_error_message("Please select an order to print.")

    def print_document(self, printer, order_data):
        document = QTextDocument()
        document.setHtml(self.format_order_for_printing(order_data))
        document.print_(printer)

    def format_order_for_printing(self, order_data):
        html = f"""
        <h1>Order Details</h1>
        <table>
            <tr><td><strong>Order ID:</strong></td><td>{order_data.get('OrderId', '')}</td></tr>
            <tr><td><strong>Seller:</strong></td><td>{order_data.get('SellerName', '')}</td></tr>
            <tr><td><strong>Buyer:</strong></td><td>{order_data.get('BuyerName', '')}</td></tr>
            <tr><td><strong>Product:</strong></td><td>{order_data.get('ProductName', '')}</td></tr>
            <tr><td><strong>Driver:</strong></td><td>{order_data.get('DriverName', '')}</td></tr>
            <tr><td><strong>Vehicle:</strong></td><td>{order_data.get('LicensePlateNumber', '')}</td></tr>
            <tr><td><strong>Loading Arm:</strong></td><td>{order_data.get('LoadingArmName', '')}</td></tr>
            <tr><td><strong>Storage Tank:</strong></td><td>{order_data.get('StorageTankName', '')}</td></tr>
            <tr><td><strong>Weighbridge:</strong></td><td>{order_data.get('WeighbridgeName', '')}</td></tr>
            <tr><td><strong>Planned Quantity:</strong></td><td>{order_data.get('PlannedQuantity', '')}</td></tr>
            <tr><td><strong>Actual Quantity:</strong></td><td>{order_data.get('ActualQuantity', '')}</td></tr>
            <tr><td><strong>Unit:</strong></td><td>{order_data.get('UnitName', '')}</td></tr>
            <tr><td><strong>Initial Weight:</strong></td><td>{order_data.get('InitialWeight', '')}</td></tr>
            <tr><td><strong>Final Weight:</strong></td><td>{order_data.get('FinalWeight', '')}</td></tr>
            <tr><td><strong>Checkout Weight:</strong></td><td>{order_data.get('CheckoutWeight', '')}</td></tr>
            <tr><td><strong>Order Date:</strong></td><td>{order_data.get('LoadingDate', '')}</td></tr>
            <tr><td><strong>Checkout Time:</strong></td><td>{order_data.get('LoadingTime', '')}</td></tr>
            <tr><td><strong>Status:</strong></td><td>{order_data.get('Status', '')}</td></tr>
            <tr><td><strong>Remarks:</strong></td><td>{order_data.get('Remarks', '')}</td></tr>
        </table>
        """
        return html

    def export_to_excel(self):
        if not self.is_connected:
            return
        file_name, _ = QFileDialog.getSaveFileName(self, "Export Orders", "", "Excel Files (*.xlsx)")
        if file_name:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            if sheet:
                sheet.title = "Orders"

                headers = [
                    "Order ID", "Seller", "Buyer", "Product", "Driver", "Vehicle", "Loading Arm",
                    "Storage Tank", "Weighbridge", "Planned Quantity", "Actual Quantity", "Unit",
                    "Initial Weight", "Final Weight", "Checkout Weight", "Order Date", "Checkout Time",
                    "Status", "Remarks"
                ]
                for col, header in enumerate(headers, start=1):
                    sheet.cell(row=1, column=col, value=header)

                for row, order in enumerate(order_module.orders, start=2):
                    sheet.cell(row=row, column=1, value=str(order.get('OrderId', '')))
                    sheet.cell(row=row, column=2, value=order.get('SellerName', ''))
                    sheet.cell(row=row, column=3, value=order.get('BuyerName', ''))
                    sheet.cell(row=row, column=4, value=order.get('ProductName', ''))
                    sheet.cell(row=row, column=5, value=order.get('DriverName', ''))
                    sheet.cell(row=row, column=6, value=order.get('LicensePlateNumber', ''))
                    sheet.cell(row=row, column=7, value=order.get('LoadingArmName', ''))
                    sheet.cell(row=row, column=8, value=order.get('StorageTankName', ''))
                    sheet.cell(row=row, column=9, value=order.get('WeighbridgeName', ''))
                    sheet.cell(row=row, column=10, value=str(order.get('PlannedQuantity', '')))
                    sheet.cell(row=row, column=11, value=str(order.get('ActualQuantity', '')))
                    sheet.cell(row=row, column=12, value=order.get('UnitName', ''))
                    sheet.cell(row=row, column=13, value=str(order.get('InitialWeight', '')))
                    sheet.cell(row=row, column=14, value=str(order.get('FinalWeight', '')))
                    sheet.cell(row=row, column=15, value=str(order.get('CheckoutWeight', '')))
                    sheet.cell(row=row, column=16, value=str(order.get('LoadingDate', '')))
                    sheet.cell(row=row, column=17, value=str(order.get('LoadingTime', '')))
                    sheet.cell(row=row, column=18, value=order.get('Status', ''))
                    sheet.cell(row=row, column=19, value=order.get('Remarks', ''))

                workbook.save(file_name)
                self.show_info_message(f"Orders exported to {file_name}")
            else:
                self.show_error_message("Failed to create Excel sheet")

class OrderDialog(QDialog):
    def __init__(self, parent=None, order_id: Optional[int] = None):
        super().__init__(parent)
        self.order_id = order_id
        self.setWindowTitle("Add Order" if order_id is None else "Edit Order")
        apply_styles(self)
        self.init_ui()

    def init_ui(self):
        layout = QFormLayout(self)

        self.inputs = {
            'seller': QComboBox(),
            'buyer': QComboBox(),
            'product': QComboBox(),
            'driver': QComboBox(),
            'vehicle': QComboBox(),
            'loading_arm': QComboBox(),
            'storage_tank': QComboBox(),
            'weighbridge': QComboBox(),
            'planned_quantity': QDoubleSpinBox(),
            'actual_quantity': QDoubleSpinBox(),
            'unit': QComboBox(),
            'initial_weight': QDoubleSpinBox(),
            'final_weight': QDoubleSpinBox(),
            'checkout_weight': QDoubleSpinBox(),
            'order_date': QDateTimeEdit(),
            'checkout_time': QDateTimeEdit(),
            'status': QComboBox(),
            'remarks': QTextEdit()
        }

        for key, widget in self.inputs.items():
            if isinstance(widget, QDoubleSpinBox):
                
                widget.setRange(0, 1000000)
                widget.setDecimals(2)
            
                

        layout.addRow("Seller:", self.inputs['seller'])
        layout.addRow("Buyer:", self.inputs['buyer'])
        layout.addRow("Product:", self.inputs['product'])
        layout.addRow("Driver:", self.inputs['driver'])
        layout.addRow("Vehicle:", self.inputs['vehicle'])
        layout.addRow("Loading Arm:", self.inputs['loading_arm'])
        layout.addRow("Storage Tank:", self.inputs['storage_tank'])
        layout.addRow("Weighbridge:", self.inputs['weighbridge'])
        layout.addRow("Planned Quantity:", self.inputs['planned_quantity'])
        layout.addRow("Actual Quantity:", self.inputs['actual_quantity'])
        layout.addRow("Unit:", self.inputs['unit'])
        layout.addRow("Initial Weight:", self.inputs['initial_weight'])
        layout.addRow("Final Weight:", self.inputs['final_weight'])
        layout.addRow("Checkout Weight:", self.inputs['checkout_weight'])
        layout.addRow("Order Date:", self.inputs['order_date'])
        layout.addRow("Checkout Time:", self.inputs['checkout_time'])
        layout.addRow("Status:", self.inputs['status'])
        layout.addRow("Remarks:", self.inputs['remarks'])

        buttons = QHBoxLayout()
        self.save_button = QPushButton("Save")
        self.cancel_button = QPushButton("Cancel")
        buttons.addWidget(self.save_button)
        buttons.addWidget(self.cancel_button)

        

        layout.addRow(buttons)

        self.save_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

        self.populate_dropdowns()
        if self.order_id:
            self.populate_data()

    def populate_dropdowns(self):
        self.inputs['seller'].addItems(item['PartyName'] for item in order_module.get_all_parties())
        self.inputs['buyer'].addItems(item['PartyName'] for item in order_module.get_all_parties())
        self.inputs['product'].addItems(item['ProductName'] for item in order_module.get_all_products())
        self.inputs['driver'].addItems(item['DriverName'] for item in order_module.get_all_drivers())
        self.inputs['vehicle'].addItems(item['LicensePlateNumber'] for item in order_module.get_all_vehicles())
        self.inputs['loading_arm'].addItems(item['LoadingArmName'] for item in order_module.get_all_loading_arms())
        self.inputs['storage_tank'].addItems(item['StorageTankName'] for item in order_module.get_all_storage_tanks())
        self.inputs['weighbridge'].addItems(item['WeighbridgeName'] for item in order_module.get_all_weighbridges())
        self.inputs['unit'].addItems(item['UnitName'] for item in order_module.get_all_units())
        self.inputs['status'].addItems(["Pending", "In Progress", "Completed", "Cancelled"])

    def populate_data(self):
        if self.order_id is not None:
            order_data = order_module.get_order_details(self.order_id)
        else:
            # Handle the case where self.order_id is None
            order_data = None  # or any other appropriate value
        if order_data:
            self.set_combo_value(self.inputs['seller'], order_data.get('SellerName'))
            self.set_combo_value(self.inputs['buyer'], order_data.get('BuyerName'))
            self.set_combo_value(self.inputs['product'], order_data.get('ProductName'))
            self.set_combo_value(self.inputs['driver'], order_data.get('DriverName'))
            self.set_combo_value(self.inputs['vehicle'], order_data.get('LicensePlateNumber'))
            self.set_combo_value(self.inputs['loading_arm'], order_data.get('LoadingArmName'))
            self.set_combo_value(self.inputs['storage_tank'], order_data.get('StorageTankName'))
            self.set_combo_value(self.inputs['weighbridge'], order_data.get('WeighbridgeName'))
            self.inputs['planned_quantity'].setValue(float(order_data.get('PlannedQuantity', 0)))
            self.inputs['actual_quantity'].setValue(float(order_data.get('ActualQuantity', 0)))
            self.set_combo_value(self.inputs['unit'], order_data.get('UnitName'))
            self.inputs['initial_weight'].setValue(float(order_data.get('InitialWeight', 0)))
            self.inputs['final_weight'].setValue(float(order_data.get('FinalWeight', 0)))
            self.inputs['checkout_weight'].setValue(float(order_data.get('CheckoutWeight', 0)))
            self.inputs['order_date'].setDateTime(order_data.get('LoadingDate'))
            self.inputs['checkout_time'].setDateTime(order_data.get('LoadingTime'))
            self.inputs['status'].setCurrentText(order_data.get('Status', ''))
            self.inputs['remarks'].setPlainText(order_data.get('Remarks', ''))

    def set_combo_value(self, combo: QComboBox, value):
        index = combo.findText(value)
        if index >= 0:
            combo.setCurrentIndex(index)

    def get_data(self) -> Dict[str, Any]:
        return {
            'OrderId': self.order_id,
            'SellerId': self.get_id_from_name(order_module.get_all_parties(), self.inputs['seller'].currentText()),
            'BuyerId': self.get_id_from_name(order_module.get_all_parties(), self.inputs['buyer'].currentText()),
            'ProductId': self.get_id_from_name(order_module.get_all_products(), self.inputs['product'].currentText()),
            'DriverId': self.get_id_from_name(order_module.get_all_drivers(), self.inputs['driver'].currentText()),
            'VehicleId': self.get_id_from_name(order_module.get_all_vehicles(), self.inputs['vehicle'].currentText()),
            'LoadingArmId': self.get_id_from_name(order_module.get_all_loading_arms(), self.inputs['loading_arm'].currentText()),
            'StorageTankId': self.get_id_from_name(order_module.get_all_storage_tanks(), self.inputs['storage_tank'].currentText()),
            'WeighbridgeId': self.get_id_from_name(order_module.get_all_weighbridges(), self.inputs['weighbridge'].currentText()),
            'PlannedQuantity': self.inputs['planned_quantity'].value(),
            'ActualQuantity': self.inputs['actual_quantity'].value(),
            'UnitId': self.get_id_from_name(order_module.get_all_units(), self.inputs['unit'].currentText()),
            'InitialWeight': self.inputs['initial_weight'].value(),
            'FinalWeight': self.inputs['final_weight'].value(),
            'CheckoutWeight': self.inputs['checkout_weight'].value(),
            'LoadingDate': self.inputs['order_date'].dateTime().toPython(),
            'LoadingTime': self.inputs['checkout_time'].dateTime().toPython(),
            'Status': self.inputs['status'].currentText(),
            'Remarks': self.inputs['remarks'].toPlainText()
        }

    def get_id_from_name(self, items, name):
        for item in items:
            if item.get('PartyName') == name or item.get('ProductName') == name or item.get('DriverName') == name or \
               item.get('LicensePlateNumber') == name or item.get('LoadingArmName') == name or \
               item.get('StorageTankName') == name or item.get('WeighbridgeName') == name or item.get('UnitName') == name:
                return item.get('PartyId') or item.get('ProductId') or item.get('DriverId') or \
                       item.get('VehicleId') or item.get('LoadingArmId') or item.get('StorageTankId') or \
                       item.get('WeighbridgeId') or item.get('UnitId')
        return None
    
    